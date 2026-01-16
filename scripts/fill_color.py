import openpyxl
from openpyxl.styles import PatternFill


# 1. Definiamo i due tipi di riempimento
def fill_cells(end_col,end_row,file):
    bianco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    grigio_chiaro = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    for l in range(2,end_row+1,1):
        for i in range(1,end_col+1,1):
            if l%2==0:
                file.cell(row=l,column=i).fill=grigio_chiaro
            else:
                file.cell(row=l,column=i).fill=bianco
    return file
