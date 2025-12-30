import asyncio
import os
import coc
from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter

# --- CONFIGURAZIONE ---
COC_EMAIL = os.getenv("COC_EMAIL")
COC_PASSWORD = os.getenv("COC_PASSWORD")
CLAN_TAG = os.getenv("CLAN_TAG")
 

def copia_stile(cella_origine, cella_destinazione):
    """
    Copia lo stile (font, bordo, riempimento, allineamento, formato numeri)
    da una cella all'altra.
    """
    if cella_origine.has_style:
        cella_destinazione.font = copy(cella_origine.font)
        cella_destinazione.border = copy(cella_origine.border)
        cella_destinazione.fill = copy(cella_origine.fill)
        cella_destinazione.number_format = copy(cella_origine.number_format)
        cella_destinazione.protection = copy(cella_origine.protection)
        cella_destinazione.alignment = copy(cella_origine.alignment)


    

async def aggiorna_membri(client,res,nome_file_input,nome_file_output):
    # Carichiamo il workbook
    print(f"Caricamento file: {nome_file_input}...")
    wb = openpyxl.load_workbook(nome_file_input)
    
    
    # --- MODIFICA SHEET 1 ---
    # Supponiamo che il primo sheet si chiami "Foglio1" (o usa wb.sheetnames[0])
    ws1 = wb['wars'] 
    
    members= await client.get_members(CLAN_TAG)
  
    nomi_members=[mem.name for mem in members]
    list_names=[]
    
   
    for idx in range( ws1.max_row,1,-1):
        val=ws1.cell(row=idx,column=1).value
        if val not in nomi_members:
            ws1.delete_rows(idx=idx, amount=1)
        else:
            list_names.append(val)
    
        
    set_1=set(nomi_members)
    set_2=set(list_names)

    ris=set_1-set_2

    old_max_row=ws1.max_row

    i=0

    
    
    for name in ris:
        current_row = old_max_row + 1 + i     
        #formula1 = f"=AVERAGE(F{current_row}:{ws1.max_column}{current_row})"
        #formula2= f"=SUM(F{current_row}:{ws1.max_column}{current_row})"
        ws1.append([
            name,
            0,
            0,
            0,
            0
        ])
        for col in range(1, ws1.max_column + 1):
            source_cell = ws1.cell(row=old_max_row, column=col)
            target_cell = ws1.cell(row=current_row, column=col)
            
            # Copiamo lo stile
            copia_stile(source_cell, target_cell)
        i += 1
    

   
    res_dict = {mem["name"]: mem for mem in res}

    index = 7 
    for i in range(7, ws1.max_column + 1):
        cell_value = ws1.cell(row=1, column=i).value
        if cell_value is None or not str(cell_value).startswith("WAR"):
            index = i
            break


    war_number = int((index - 5) / 2)
    ws1.cell(row=1, column=index).value = f"WAR-{war_number} (1)"
    ws1.cell(row=1, column=index + 1).value = f"WAR-{war_number} (2)"

    for idx in range(2, ws1.max_row + 1):
        temp_name = ws1.cell(row=idx, column=1).value
        member_data = res_dict.get(temp_name)
        if member_data:
            if member_data.get("war_skip"):
                ws1.cell(row=idx, column=3).value = (ws1.cell(row=idx, column=3).value or 0) + 1
            ws1.cell(row=idx, column=2).value = (ws1.cell(row=idx, column=2).value or 0) + member_data.get("atk_skip", 0)
            ws1.cell(row=idx, column=4).value =(ws1.cell(row=idx, column=2).value or 0) +1
     
            ws1.cell(row=idx, column=index).value = member_data.get("atk1")
            ws1.cell(row=idx, column=index+1).value = member_data.get("atk2")
            

            last_col = get_column_letter(index + 1)

            ws1.cell(row=idx, column=5).value = f"=SUM(G{idx}:{last_col}{idx})"
            ws1.cell(row=idx, column=6).value = f"=ROUND(AVERAGE(G{idx}:{last_col}{idx}),1)"


    wb.save(nome_file_output)



async def esporta_dati():
    # 1. Inizializzazione Client
    async with coc.Client(key_names="PC_Locale_Key") as client:
        try:
            await client.login(COC_EMAIL, COC_PASSWORD)
            members= await client.get_members(CLAN_TAG)
            


            # 2. Estrazione Membri del Clan
            print(f"Recupero war per il clan: {CLAN_TAG}...")
            clan_war = await client.get_current_war(CLAN_TAG)
            
            # 3. Recupero parallelo dei dati (Sostituisce ThreadPoolExecutor)
            print(f"E'una CWL: {clan_war.is_cwl} ")
            print(f"stato war: {clan_war.state} ")

            results=[]
            partecipants=clan_war.members
            for part in partecipants:
                if part.name in [mem.name for mem in members]:
                    list_star=[]
                    if part.attacks:
                        for atk in part.attacks:
                            #print(atk.stars)
                            list_star.append(atk.stars)
                    if len(list_star)==0:
                        #print(f"{part.name}:number of attacks:{len(part.attacks)} no attacks")
                        results.append({"name":part.name, "war_skip":True,"atk_skip":2,"atk1":"SKIP","atk2":"SKIP"})
                    
                    elif len(list_star)==1:
                        #print(f"{part.name}:number of attacks:{len(part.attacks)} atk1:{list_star[0]} ")
                        results.append({"name":part.name, "war_skip":False,"atk_skip":1,"atk1":list_star[0],"atk2":"SKIP"})
                    
                    else:
                        #print(f"{part.name}:number of attacks:{len(part.attacks)} atk1:{list_star[0]} atk2:{list_star[1]}")
                        results.append({"name":part.name, "war_skip":False,"atk_skip":0,"atk1":list_star[0],"atk2":list_star[1]})

            await aggiorna_membri(client,results,"rewards.xlsx","rewards.xlsx")
           

                
                

        except coc.errors.InvalidCredentials:
            print("Credenziali non valide!")
            return 
        except Exception as e:
            print(f"Errore generale: {e}")
            return 
async def main():
    # 1. Recupera i dati (DEVI USARE await)
    print("Inizio recupero dati dai server Supercell...")
    await esporta_dati()  
    
    

if __name__ == "__main__":
    # Avvia tutto il ciclo
    asyncio.run(main())




