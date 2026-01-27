import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
from datetime import date, timedelta,datetime
from points import esporta_dati
from fill_color import fill_cells
import asyncio


def copia_stile(cella_origine, cella_destinazione):
    """
    Copia lo stile (font, bordo, riempimento, allineamento, formato numeri)
    da una cella all'altra.
    """
    if cella_origine.has_style:
        cella_destinazione.font = copy(cella_origine.font)
        cella_destinazione.border = copy(cella_origine.border)
        cella_destinazione.fill = copy(cella_origine.fill)
        cella_destinazione.number_format = copy(cella_origine.number_format)
        cella_destinazione.protection = copy(cella_origine.protection)
        cella_destinazione.alignment = copy(cella_origine.alignment)

def modifica_excel(nome_file_input, nome_file_output,members):
    # Carichiamo il workbook
    print(f"Caricamento file: {nome_file_input}...")
    wb = openpyxl.load_workbook(nome_file_input)
    
    
    # --- MODIFICA SHEET 1 ---
    # Supponiamo che il primo sheet si chiami "Foglio1" (o usa wb.sheetnames[0])
    ws1 = wb['clan games'] 
    nomi_members = [m['name'] for m in members]
    
    list_names=[]

   
    for idx in range( ws1.max_row,1,-1):
        val=ws1.cell(row=idx,column=1).value
        if val not in nomi_members:
            ws1.delete_rows(idx=idx, amount=1)
        else:
            list_names.append(val)
    
        
    set_1=set(nomi_members)
    set_2=set(list_names)

    ris=set_1-set_2

    old_max_row=ws1.max_row

    i=0
    x = date.today()
    
    for name in ris:
        current_row = old_max_row + 1 + i     
        #formula1 = f"=AVERAGE(F{current_row}:{ws1.max_column}{current_row})"
        #formula2= f"=SUM(F{current_row}:{ws1.max_column}{current_row})"
        ws1.append([
            name, 
            x.strftime("%d/%m/%Y") 
        ])
        for col in range(1, ws1.max_column + 1):
            source_cell = ws1.cell(row=old_max_row, column=col)
            target_cell = ws1.cell(row=current_row, column=col)
            
            # Copiamo lo stile
            copia_stile(source_cell, target_cell)
        i += 1
    y=ws1.max_column+1
    
    lett=get_column_letter(y) 
    old_points=[]
    for row in range(2,ws1.max_row+1):
        ws1.cell(row=row, column=3).value = f"=ROUND(AVERAGE(F{row}:{lett}{row}), 0)"
        ws1.cell(row=row, column=4).value = f"=SUM(F{row}:{lett}{row})" 
        if(ws1.cell(row=row,column=5).value is None):
            name_to_search=ws1.cell(row=row,column=1).value
            for mem in members:
                if(mem["name"]==name_to_search):
                    ws1.cell(row=row,column=5).value=mem["points"]
        old_points.append({"name":ws1.cell(row=row,column=1).value,"points": ws1.cell(row=row,column=5).value,"index": row })
    today=date.today()
    old_points.sort(key=lambda x: x["name"])
    members.sort(key=lambda y: y["name"])

    last_col=6
    if(6<=ws1.max_column):
        for i in range(6,ws1.max_column + 1,1):
            current=ws1.cell(row=2,column=i).value
            next_cell=ws1.cell(row=2,column=(i+1)).value
            if(isinstance(current,int) and next_cell is None):
                last_col=i+1
                break
            elif(current is None and i==ws1.max_column):
                print("documento vuoto")
    nextCG_start1=ws1.cell(row=1,column=last_col).value
    lastCG_start1=ws1.cell(row=1,column=last_col-1).value
    
    lastCG_start = datetime.strptime(lastCG_start1, "%d/%m/%Y").date()
    lastCG_end=lastCG_start + timedelta(days=6)
    nextCG_start = datetime.strptime(nextCG_start1, "%d/%m/%Y").date()
    nextCG_end= nextCG_start + timedelta(days=6)
    
   


    
    if (today<nextCG_start):
        for i in range(0,ws1.max_row-1):
            if(old_points[i]["name"]!=members[i]["name"]):
                print("ERRORE NELL'ASSEGNAZIONE DEI PUNTI!")
            else:
                point_var=ws1.cell(row=old_points[i]["index"],column=last_col-1).value
                test_data=ws1.cell(row=old_points[i]["index"],column=2).value
                ach=ws1.cell(row=old_points[i]["index"],column=5).value
                
                
                if point_var is None and datetime.strptime((test_data), "%d/%m/%Y").date()<= lastCG_end :
                    ws1.cell(row=old_points[i]["index"],column=last_col-1).value=0
                    ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
                elif(point_var is not None):
                    ws1.cell(row=old_points[i]["index"],column=last_col-1).value+=(members[i]["points"]-ach)
                    if (ws1.cell(row=old_points[i]["index"],column=last_col-1).value or 0 )>10000:
                        ws1.cell(row=old_points[i]["index"],column=last_col-1).value=10000
                    #if members[i]["points"]-ach!=0:
                        #print(f"{members[i]['name']}:{members[i]['points']-ach}")
                    ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
                    
    else:
        if today>nextCG_end:
            for i in range(0,ws1.max_row-1):
                if(old_points[i]["name"]!=members[i]["name"]):
                    print("ERRORE NELL'ASSEGNAZIONE DEI PUNTI!")    
                else:
                    point_var=ws1.cell(row=old_points[i]["index"],column=last_col).value
                    ach=ws1.cell(row=old_points[i]["index"],column=5).value
                    if(point_var is None and datetime.strptime((ws1.cell(row=old_points[i]["index"],column=2).value), "%d/%m/%Y").date()<= nextCG_end ):
                        ws1.cell(row=old_points[i]["index"],column=last_col).value=0+(members[i]["points"]-ach)  
                        if (ws1.cell(row=old_points[i]["index"],column=last_col).value or 0 )>10000:
                            ws1.cell(row=old_points[i]["index"],column=last_col).value=10000
                        #if members[i]["points"]-ach!=0:
                            #print(f"{members[i]['name']}:{members[i]['points']-ach}")
                        ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
        else:
            if today.day>=22 and today.day<29:
                for i in range(0,ws1.max_row-1):
                    if(old_points[i]["name"]!=members[i]["name"]):
                        print("ERRORE NELL'ASSEGNAZIONE DEI PUNTI!")    
                    else:
                        point_var=ws1.cell(row=old_points[i]["index"],column=last_col).value
                        ach=ws1.cell(row=old_points[i]["index"],column=5).value
                        if(point_var is None and datetime.strptime((ws1.cell(row=old_points[i]["index"],column=2).value), "%d/%m/%Y").date()<= nextCG_end ):
                            ws1.cell(row=old_points[i]["index"],column=last_col).value=0+(members[i]["points"]-ach) 
                            if (ws1.cell(row=old_points[i]["index"],column=last_col).value or 0 )>10000:
                                ws1.cell(row=old_points[i]["index"],column=last_col).value=10000
                            #if members[i]["points"]-ach!=0:
                                #print(f"{members[i]['name']}:{members[i]['points']-ach}")
                            ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
            else:
                for i in range(0,ws1.max_row-1):
                    if(old_points[i]["name"]!=members[i]["name"]):
                        print("ERRORE NELL'ASSEGNAZIONE DEI PUNTI!")    
                    else:
                        point_var=ws1.cell(row=old_points[i]["index"],column=last_col-1).value
                        ach=ws1.cell(row=old_points[i]["index"],column=5).value
                        if(point_var is None and datetime.strptime((ws1.cell(row=old_points[i]["index"],column=2).value), "%d/%m/%Y").date()<= lastCG_end ):
                            ws1.cell(row=old_points[i]["index"],column=last_col-1).value=0+(members[i]["points"]-ach) 
                            if (ws1.cell(row=old_points[i]["index"],column=last_col-1).value or 0 )>10000:
                                ws1.cell(row=old_points[i]["index"],column=last_col-1).value=10000
                            #if members[i]["points"]-ach!=0:
                                #print(f"{members[i]['name']}:{members[i]['points']-ach}")
                            ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
                        elif( datetime.strptime((ws1.cell(row=old_points[i]["index"],column=2).value), "%d/%m/%Y").date()<= lastCG_end):
                            ws1.cell(row=old_points[i]["index"],column=last_col-1).value=(ws1.cell(row=old_points[i]["index"],column=last_col-1).value or 0) + (members[i]["points"]-ach)
                            if (ws1.cell(row=old_points[i]["index"],column=last_col-1).value or 0 )>10000:
                                ws1.cell(row=old_points[i]["index"],column=last_col-1).value=10000
                           # if members[i]["points"]-ach!=0:
                                #print(f"{members[i]['name']}:{members[i]['points']-ach}")
                            ws1.cell(row=old_points[i]["index"],column=5).value=members[i]["points"]
    
    print(f"Aggiornati punti per {len(members)} membri nella colonna {last_col}")                
                




    

    # --- SALVATAGGIO ---
    #ws1=fill_cells(ws1.max_column,ws1.max_row,ws1)
    print(f"Salvataggio file: {nome_file_output}...")
    wb.save(nome_file_output)
    


async def main():
    # 1. Recupera i dati (DEVI USARE await)
    print("Inizio recupero dati dai server Supercell...")
    players = await esporta_dati()  # Qui serviva l'await!
    
    # 2. Passa i dati alla funzione Excel
    # Se modifica_excel non è async, la chiami normalmente
    if players:
        modifica_excel('cc_cg_events.xlsx', 'cc_cg_events.xlsx', players)
        print("Processo completato!")
    else:
        print("Nessun dato recuperato.")

if __name__ == "__main__":
    # Avvia tutto il ciclo
    asyncio.run(main())

