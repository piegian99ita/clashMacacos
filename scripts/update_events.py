import gspread
import openpyxl
from oauth2client.service_account import ServiceAccountCredentials
import json
import os

# --- FUNZIONE HELPER PER DETERMINARE IL SEPARATORE ---
def get_formula_separator(spreadsheet_locale):
    """
    Restituisce ';' se il locale usa il punto e virgola (Europa/Sud America),
    altrimenti restituisce ',' (USA/UK/Standard).
    """
    # Lista parziale di locali che usano il punto e virgola
    # Italia, Germania, Francia, Spagna, Brasile, Olanda, Russia, ecc.
    semicolon_locales = ['it_', 'de_', 'fr_', 'es_', 'pt_', 'nl_', 'ru_', 'pl_']
    
    # Se il locale del foglio inizia con uno di questi prefissi
    for prefix in semicolon_locales:
        if spreadsheet_locale.startswith(prefix):
            return ';'
    
    return ','
# -----------------------------------------------------

# 1. Configurazione Credenziali
json_creds = json.loads(os.environ['GCP_SERVICE_ACCOUNT_KEY'])
scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
creds = ServiceAccountCredentials.from_json_keyfile_dict(json_creds, scope)
client = gspread.authorize(creds)

# 2. Apri il foglio Google e LEGGI IL LOCALE
spreadsheet_id = os.environ['SPREADSHEET_ID_EVENTS']
spreadsheet = client.open_by_key(spreadsheet_id)

# Otteniamo la lingua impostata nel foglio (es. 'it_IT' o 'en_US')
current_locale = spreadsheet.locale
print(f"Lingua rilevata nel Google Sheet: {current_locale}")

# Determiniamo quale separatore usare
target_separator = get_formula_separator(current_locale)
print(f"Separatore formule selezionato: '{target_separator}'")

# 3. Impostazioni file Excel
file_path = 'cc_cg_events.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=False)

for sheet_name in wb.sheetnames:
    print(f"Elaborazione foglio: {sheet_name}...")
    
    ws = wb[sheet_name]
    data_to_upload = []
    
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            val = cell.value
            
            # --- LOGICA DI CONVERSIONE INTELLIGENTE ---
            if isinstance(val, str) and val.startswith('='):
                # Le formule generate da Python/Excel sono solitamente in formato USA (con virgola)
                # Se il target richiede ';', sostituiamo le virgole
                if target_separator == ';':
                     val = val.replace(',', ';')
                # Se il target richiede ',' (USA), lasciamo così com'è 
                # (assumendo che l'Excel generato sia già in formato US)
            
            if val is None:
                val = ""
            
            row_data.append(val)
        data_to_upload.append(row_data)

    # 4. Caricamento su Google Sheets
    try:
        worksheet = spreadsheet.worksheet(sheet_name)
    except gspread.exceptions.WorksheetNotFound:
        worksheet = spreadsheet.add_worksheet(title=sheet_name, rows="100", cols="20")
        print(f"Creata nuova tab: {sheet_name}")

    worksheet.clear()
    
    spreadsheet.values_update(
        f"'{sheet_name}'!A1",
        params={'valueInputOption': 'USER_ENTERED'},
        body={'values': data_to_upload}
    )
    
    print(f"Tab '{sheet_name}' aggiornata.")

print("Aggiornamento completo riuscito.")