import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from datetime import date
import asyncio
import os
import coc

# --- CONFIGURAZIONE ---
# Leggiamo i dati dai Secrets di GitHub
COC_EMAIL = os.getenv("COC_EMAIL")
COC_PASSWORD = os.getenv("COC_PASSWORD")
CLAN_TAG = os.getenv("CLAN_TAG")
FILE_EXCEL = "cc_cg_events.xlsx"

def copia_stile(cella_origine, cella_destinazione):
    if cella_origine.has_style:
        cella_destinazione.font = copy(cella_origine.font)
        cella_destinazione.border = copy(cella_origine.border)
        cella_destinazione.fill = copy(cella_origine.fill)
        cella_destinazione.number_format = copy(cella_origine.number_format)
        cella_destinazione.alignment = copy(cella_origine.alignment)

def aggiorna_struttura_membri(wb, lista_nomi_api):
    ws = wb['CLAN CAPITAL']
    for idx in range(ws.max_row, 1, -1):
        nome_excel = ws.cell(row=idx, column=1).value
        if nome_excel and nome_excel not in lista_nomi_api:
            ws.delete_rows(idx=idx, amount=1)

    nomi_gia_presenti = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    nuovi_membri = [n for n in lista_nomi_api if n not in nomi_gia_presenti]

    for nome in nuovi_membri:
        prossima_riga = ws.max_row + 1
        ws.cell(row=prossima_riga, column=1).value = nome
        ws.cell(row=prossima_riga, column=2).value = date.today().strftime("%d/%m/%Y")
        if prossima_riga > 2:
            for col in range(1, ws.max_column + 1):
                copia_stile(ws.cell(row=prossima_riga-1, column=col), ws.cell(row=prossima_riga, column=col))
    return ws

async def main():
    if not COC_EMAIL or not COC_PASSWORD:
        print("Errore: Credenziali mancanti nelle variabili d'ambiente.")
        return

    async with coc.Client(key_names="GitHub_Actions_Key") as client:
        try:
            await client.login(COC_EMAIL, COC_PASSWORD)
            
            if not os.path.exists(FILE_EXCEL):
                print(f"Errore: {FILE_EXCEL} non trovato nel repository.")
                return

            wb = openpyxl.load_workbook(FILE_EXCEL)
            clan = await client.get_clan(CLAN_TAG)
            lista_nomi = [m.name for m in clan.members]
            ws = aggiorna_struttura_membri(wb, lista_nomi)

            raid_log = await client.get_raid_log(CLAN_TAG)
            raid_attuale = next((r for r in raid_log if str(r.state).lower() == 'ongoing'), None)
            if raid_attuale:
                start_dt = raid_attuale.start_time.time
                end_dt=raid_attuale.end_time.time
                if raid_attuale:
                    data_inizio_str = raid_attuale.start_time.date().strftime("%d/%m/%Y")
                    colonna_target = None
                    for col in range(4, ws.max_column + 1):
                        if ws.cell(row=1, column=col).value.startswith("Colonna"):
                            colonna_target = col
                            if date.today()>=start_dt.date()  and date.today()<=end_dt.date():
                                ws.cell(row=1, column=col).value=data_inizio_str
                            
                            break
                    
                    if not colonna_target:
                        colonna_target = 4
                        while ws.cell(row=1, column=colonna_target).value is not None:
                            colonna_target += 1
                        ws.cell(row=1, column=colonna_target).value = data_inizio_str

                    membri_raid = {m.name: m.attack_count for m in raid_attuale.members}
                    for i in range(2, ws.max_row + 1):
                        nome = ws.cell(row=i, column=1).value
                        ws.cell(row=i, column=colonna_target).value = membri_raid.get(nome, 0)

                    ultima_col_lettera = get_column_letter(ws.max_column)
                    for row in range(2, ws.max_row + 1):
                        # Arrotondamento aggiunto come richiesto
                        ws.cell(row=row, column=3).value = f"=ROUND(AVERAGE(D{row}:{ultima_col_lettera}{row}), 0)"
                        if ws.cell(row=row,column=colonna_target).value is None:
                            ws.cell(row=row,column=colonna_target).value=0
            else:
                print("NESSUN RAID IN CORSO")   
                ultima_col_lettera = get_column_letter(ws.max_column) 
                for row in range(2, ws.max_row + 1):
                    # Arrotondamento aggiunto come richiesto
                    ws.cell(row=row, column=3).value = f"=ROUND(AVERAGE(D{row}:{ultima_col_lettera}{row}), 0)"
                    

            wb.save(FILE_EXCEL)
            print("Excel aggiornato con successo.")

        except Exception as e:
            print(f"Errore: {e}")

if __name__ == "__main__":
    asyncio.run(main())