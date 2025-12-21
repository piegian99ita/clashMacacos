import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
import datetime
from datetime import date, datetime

import json
import requests

def copia_stile(cella_origine, cella_destinazione):
    """
    Copia lo stile (font, bordo, riempimento, allineamento, formato numeri)
    da una cella all'altra.
    """
    if cella_origine.has_style:
        cella_destinazione.font = copy(cella_origine.font)
        cella_destinazione.border = copy(cella_origine.border)
        cella_destinazione.fill = copy(cella_origine.fill)
        cella_destinazione.number_format = copy(cella_origine.number_format)
        cella_destinazione.protection = copy(cella_origine.protection)
        cella_destinazione.alignment = copy(cella_origine.alignment)



def modifica_excel(nome_file_input, nome_file_output,members):
    # Carichiamo il workbook
    print(f"Caricamento file: {nome_file_input}...")
    wb = openpyxl.load_workbook(nome_file_input)
    ws2 = wb['CLAN CAPITAL']
    
    # --- MODIFICA SHEET 1 ---
    # Supponiamo che il primo sheet si chiami "Foglio1" (o usa wb.sheetnames[0])
    list_names2=[]
    for idx in range( ws2.max_row,1,-1):
        val=ws2.cell(row=idx,column=1).value
        if val not in members:
            ws2.delete_rows(idx=idx, amount=1)
        else:
            list_names2.append(val)

        
    set_1=set(members)
    
    set_3=set(list_names2)
    
    ris2=set_1-set_3
    
    old_max_row2=ws2.max_row
    x = date.today()
    l=0
    for name in ris2:
        current_row = old_max_row2 + 1 + l   
        ws2.append([
            name, 
            x.strftime("%d/%m/%Y") 
        ])
        for col in range(1, ws2.max_column):
            source_cell = ws2.cell(row=old_max_row2, column=col)
            target_cell = ws2.cell(row=current_row, column=col)
            
            # Copiamo lo stile
            copia_stile(source_cell, target_cell)
        l += 1
    
    y=ws2.max_column
    lett=get_column_letter(y+1)    
    
    for row in range(2,ws2.max_row+1):
        ws2.cell(row=row, column=3).value = f"=AVERAGE(D{row}:{lett}{row})" 
    
    # --- SALVATAGGIO ---
    print(f"Salvataggio file: {nome_file_output}...")
    return wb
    


nome_file="./secret.json"

with open(nome_file, 'r', encoding='utf-8') as file:
        # Usa json.load per convertire il JSON (testo) in un oggetto Python (Dizionario o Lista)
        dati = json.load(file)


token=dati['token']
clan_tag=dati['clan_tag']

HEADERS={
    'Authorization': f'Bearer {token}'
}

URL_API="https://api.clashofclans.com/v1/clans/%23"+clan_tag+"/members"
risposta_coc = requests.get(URL_API, headers=HEADERS)

try:
    risposta_coc.raise_for_status()
    dati_clan = risposta_coc.json()
    
    lista_membri = dati_clan.get('items', [])
    print("\n--- Lista Membri Estratta ---")
    #print(f"Membri trovati: {len(lista_membri)}")

    if lista_membri:
        # Stampa i dati del primo membro come esempio
        primo_membro = lista_membri[0]
        lista_membri2=[]
        for members in lista_membri:
            lista_membri2.append(members.get('name'))
            #print(members.get('name'))







except requests.exceptions.HTTPError as e:
    print(f"\nErrore nella richiesta API di Clash of Clans: {e}")
    # Stampa la risposta testuale per vedere il messaggio di errore esatto (es. IP non autorizzato)
    print("Messaggio dal server:", risposta_coc.text)

wb=modifica_excel('prova.xlsx', 'prova.xlsx',lista_membri2)

ws2 = wb['CLAN CAPITAL']
print(ws2.max_column)
last_col=4
if(4<=ws2.max_column):
    for i in range(4,ws2.max_column + 1,1):
        current=ws2.cell(row=2,column=i).value
        next_cell=ws2.cell(row=2,column=(i+1)).value
        if(isinstance(current,int) and next_cell is None):
            last_col=i+1
            break
        elif(current is None and i==ws2.max_column):
            print("documento vuoto")



URL_API2="https://api.clashofclans.com/v1/clans/%23"+clan_tag+"/capitalraidseasons"
risposta_coc2 = requests.get(URL_API2, headers=HEADERS)
on=False
try:
    risposta_coc2.raise_for_status()
    dati_clan = risposta_coc2.json()
    
    raid_log = dati_clan.get('items', [])
    ongoing_raids = [item for item in raid_log if item.get('state') == 'ongoing']
    # 2. Controlliamo se esiste almeno un raid attivo
    if ongoing_raids:
        # Prendiamo il primo (e solitamente unico) raid 'ongoing'
        on=True
        current_raid = ongoing_raids[0]


        on=True
        current_raid = ongoing_raids[0]
        
        # 3. Estraiamo la lista dei membri
        startTime = current_raid.get('startTime')
        data_pulita = startTime.replace("Z", "+00:00")
        startDate = datetime.fromisoformat(data_pulita).date()
        
        endTime = current_raid.get('endTime')
        data_pulita2 = endTime.replace("Z", "+00:00")
        endDate = datetime.fromisoformat(data_pulita2).date()
        
        lastWar=ws2.cell(row=1,column=last_col-1).value
        dataWar = datetime.strptime(lastWar, "%d/%m/%Y").date()

        if(dataWar>=startDate and dataWar<endDate):
            
            # 3. Estraiamo la lista dei membri
            members = current_raid.get('members', [])
            last_col=last_col-1
            for i in range(2, ws2.max_row + 1):
                # Leggiamo il nome dalla colonna 1 di Excel
                nome_excel = ws2.cell(row=i, column=1).value
                
                
                # Cerchiamo questo nome nella lista di dizionari 'members'
                # Usiamo next() per trovare il primo membro che corrisponde
                member_found = next((m for m in members if m.get('name') == nome_excel), None)
                
                if member_found:
                    attacchi = member_found.get('attacks', 0)
                    # Aggiorniamo la colonna 2 con gli attacchi trovati
                    ws2.cell(row=i, column=last_col).value = attacchi
                    print(f"Aggiornato {nome_excel}: {attacchi} attacchi")
                else:
                    print(f"Il giocatore {nome_excel} non ha partecipato al raid.")
                    if(nome_excel is not None):
                        ws2.cell(row=i, column=last_col).value = 0


        else:
            # 3. Estraiamo la lista dei membri
            members = current_raid.get('members', [])
            ws2.cell(row=1,column=last_col).value=startDate.strftime("%d/%m/%Y")
            for i in range(2, ws2.max_row + 1):
                # Leggiamo il nome dalla colonna 1 di Excel
                nome_excel = ws2.cell(row=i, column=1).value
                
                
                # Cerchiamo questo nome nella lista di dizionari 'members'
                # Usiamo next() per trovare il primo membro che corrisponde
                member_found = next((m for m in members if m.get('name') == nome_excel), None)
                
                if member_found:
                    attacchi = member_found.get('attacks', 0)
                    # Aggiorniamo la colonna 2 con gli attacchi trovati
                    ws2.cell(row=i, column=last_col).value = attacchi
                    print(f"Aggiornato {nome_excel}: {attacchi} attacchi")
                else:
                    print(f"Il giocatore {nome_excel} non ha partecipato al raid.")
                    ws2.cell(row=i, column=last_col).value = 0
        
        
        
       
        wb.save('prova.xlsx')
    else:
        print("Nessun weekend di assalti in corso al momento.")
        wb.save('prova.xlsx')
    







except requests.exceptions.HTTPError as e:
    print(f"\nErrore nella richiesta API di Clash of Clans: {e}")
    # Stampa la risposta testuale per vedere il messaggio di errore esatto (es. IP non autorizzato)
    print("Messaggio dal server:", risposta_coc.text)




